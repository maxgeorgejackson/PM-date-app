import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import argparse

def parse_date(date_val):
    if pd.isna(date_val):
        return None
    if isinstance(date_val, pd.Timestamp):
        return date_val.to_pydatetime()
    if isinstance(date_val, datetime):
        return date_val
    date_str = str(date_val)
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d%m%y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def extract_dates(text):
    if pd.isna(text):
        return []
    pattern = r'\b(\d{2}/\d{2}/\d{4}|\d{2}/\d{2}/\d{2}|\d{6})\b'
    return re.findall(pattern, text)

def main(input_file):
    df = pd.read_excel(input_file)

    results = []
    all_weeks = []

    for idx, row in df.iterrows():
        pm_date = parse_date(row['Date of PM'])
        mets_date = parse_date(row['Mets Development'])
        last_followup_date = parse_date(row['Date of last follow up/death'])

        mets_diff = (mets_date - pm_date).days // 7 + 1 if pm_date and mets_date else None
        last_followup_diff = (last_followup_date - pm_date).days // 7 + 1 if pm_date and last_followup_date else None

        whole_blood_dates = [parse_date(d) for d in extract_dates(str(row['Whole blood']))]
        followup_blood_dates = [parse_date(d) for d in extract_dates(str(row['Follow up bloods']))]

        whole_blood_weeks = [(d - pm_date).days // 7 + 1 for d in whole_blood_dates if d and pm_date]
        followup_blood_weeks = [(d - pm_date).days // 7 + 1 for d in followup_blood_dates if d and pm_date]

        for w in [mets_diff, last_followup_diff]:
            if w:
                all_weeks.append(w)
        all_weeks.extend(whole_blood_weeks)
        all_weeks.extend(followup_blood_weeks)

        results.append({
            'Biobank Number': row['Biobank Number'],
            'Mets Weeks': mets_diff,
            'Date of last follow up/death Weeks': last_followup_diff,
            'Whole Blood Weeks': whole_blood_weeks,
            'Follow-up Blood Weeks': followup_blood_weeks
        })

    max_week = max(all_weeks) if all_weeks else 1

    wb = load_workbook(input_file)

    if 'Weekly Events' in wb.sheetnames:
        ws = wb['Weekly Events']
        wb.remove(ws)
    ws = wb.create_sheet('Weekly Events')

    color_map = {
        'Mets Weeks': 'FFC7CE',                      # Light Red
        'Date of last follow up/death Weeks': 'C6EFCE', # Light Green
        'Whole Blood Weeks': 'FFEB9C',                # Light Yellow
        'Follow-up Blood Weeks': 'BDD7EE'             # Light Blue
    }

    fill_map = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') for k,v in color_map.items()}

    ws.cell(row=1, column=1, value='Biobank Number')
    for w in range(1, max_week + 1):
        ws.cell(row=1, column=w + 1, value=f'Week {w}')

    for i, r in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=r['Biobank Number'])

        for w in range(1, max_week + 1):
            cell = ws.cell(row=i, column=w + 1)
            if r['Mets Weeks'] == w:
                cell.fill = fill_map['Mets Weeks']
            elif r['Date of last follow up/death Weeks'] == w:
                cell.fill = fill_map['Date of last follow up/death Weeks']
            elif w in r['Whole Blood Weeks']:
                cell.fill = fill_map['Whole Blood Weeks']
            elif w in r['Follow-up Blood Weeks']:
                cell.fill = fill_map['Follow-up Blood Weeks']

    if 'Legend' in wb.sheetnames:
        wb.remove(wb['Legend'])
    legend_ws = wb.create_sheet('Legend')

    legend_ws.cell(row=1, column=1, value="Legend/Key")
    for i, (event, color) in enumerate(color_map.items(), start=2):
        c = legend_ws.cell(row=i, column=1, value=event)
        c.fill = fill_map[event]

    output_excel_file = input_file.replace('.xlsx', '_with_weeks_colored.xlsx')
    wb.save(output_excel_file)

    # Write out the weeks values to a txt file
    output_txt_file = input_file.replace('.xlsx', '_weeks_values.txt')
    with open(output_txt_file, 'w') as f:
        for r in results:
            f.write(f"Biobank Number: {r['Biobank Number']}\n")
            f.write(f"  PM to Met development (Weeks): {r['Mets Weeks']}\n")
            f.write(f"  PM to Date of last follow up/death (Weeks): {r['Date of last follow up/death Weeks']}\n")
            f.write(f"  PM to Whole Blood (Weeks): {', '.join(map(str, r['Whole Blood Weeks'])) if r['Whole Blood Weeks'] else 'None'}\n")
            f.write(f"  PM to Follow-up Blood(s) (Weeks): {', '.join(map(str, r['Follow-up Blood Weeks'])) if r['Follow-up Blood Weeks'] else 'None'}\n")
            f.write("\n")

    print(f"Saved weekly coloring and legend in '{output_excel_file}'.")
    print(f"Saved numeric weeks values in '{output_txt_file}'.")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Process Excel file to create week-based event coloring.")
    parser.add_argument('input_file', type=str, help='Path to input Excel file (.xlsx)')

    args = parser.parse_args()
    main(args.input_file)
